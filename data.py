from openpyxl import load_workbook
import sys
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manager

DATA_PATH = '/homeL/lin/Documents/data.xlsx'
LABEL_TICK_FONT_SIZE = 17
LABEL_FONT_SIZE = 20
LINE_WIDTH = 3


def get_group_data(ws, group_num):
    group_data = {}
    group_data_list = []
    for column in ws.iter_cols(min_row = 3, min_col = 1 + group_num*7, max_col = 6 + group_num*7, values_only = True):
        group_data_list.append(np.array(column))

    group_data['V1'] = group_data_list[0]
    group_data['A1'] = group_data_list[1]
    group_data['ohm'] = group_data_list[4][0]

    #### processing ####
    group_data['V1_'] = 0.22273 + group_data['V1'] - group_data['ohm']*group_data['A1']
    group_data['A1_'] = 1000*group_data['A1']

    for row in ws.iter_rows(min_row = 1, max_row = 1, values_only = True):
        group_data['name'] = str(row[7*group_num])


    return group_data


def plot(groups_list):
    fig = plt.figure(figsize = (8, 6))
    ax = plt.gca()

    for group in groups_list:
        x = group['V1_']
        y = group['A1_']
        # find the firt value above 1.2
        idx = np.where(x > 1.2)[0][0]
        # minus base value
        y = y-y[idx]

        ###plot####
        ax.plot(x, y, label = group['name'], linewidth = LINE_WIDTH)

    font = font_manager.FontProperties(family='Arial', weight='bold', style='normal', size=LABEL_TICK_FONT_SIZE)

    legend = ax.legend(loc='upper left', frameon = False, prop = font)
    ax.set_xlim(1.2, 1.8)
    ax.set_ylim(0, 40)
    ax.set_xlabel('Potential (V vs. RHE)', fontname = 'Arial', fontweight = 'bold', fontsize = LABEL_FONT_SIZE)
    ax.set_ylabel('Current  density (mA $\mathregular{cm^{-2}}$)', fontname = 'Arial', fontweight = 'bold', fontsize = LABEL_FONT_SIZE)
    # ax.set_xticks(np.linspace(1.2, 1.8, num = 7))
    ax.xaxis.set_major_locator(plt.MultipleLocator(base = 0.2))
    ax.xaxis.set_minor_locator(plt.MultipleLocator(base = 0.1))
    ax.yaxis.set_major_locator(plt.MultipleLocator(base = 10))
    ax.yaxis.set_minor_locator(plt.MultipleLocator(base = 5))
    for label in ax.get_xticklabels():
        label.set_fontproperties(font)

    for label in ax.get_yticklabels():
        label.set_fontproperties(font)
    #
    # ax.xaxis.set_tick_params(labelsize=15)
    # ax.yaxis.set_tick_params(labelsize=15)
    ax.tick_params(direction='out', which = 'major', length=6, width=3)
    ax.tick_params(direction='out', which = 'minor', length=4, width=3)
    # ax.set_xticklabels(labels = [' ', 1.2, 1.4, 1.6, 1.8], fontdict={'fontweight': 'bold', 'fontsize':15, 'fontname':'Arial'})
    # print(ax.spines.values())
    [i.set_linewidth(3) for i in ax.spines.values()]

    # plt.show()
    plt.savefig('experiment.eps')
    plt.show()




def main():
    ##### load data from file #####
    wb = load_workbook(DATA_PATH)
    ws = wb['Sheet1']

    ##### put data into experiment groups #####
    group_column = 6
    group_num = int((ws.max_column + 1)/(group_column + 1))
    groups_list = []

    for i in range(group_num):
        group_data= get_group_data(ws, i)
        groups_list.append(group_data)

    # print('load %d groups experiment data' % (group_num))
    # print(groups_list)

    plot(groups_list)



if __name__ == '__main__':
    main()
